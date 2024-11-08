import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from tabulate import tabulate
import pandas as pd
import datetime
import sqlite3
from sqlite3 import Error
import sys
import matplotlib.pyplot as plt
fecha_actual = datetime.date.today()

unidades={}
clientes={}
prestamos={}
colores={1:"Rojo",2:"Verde",3:"Azul",4:"Amarillo",5:"Naranja",6:"Morado"}
tabla_unidad = "UNIDADES"
tabla_cliente = "CLIENTES"

#Tabla Unidades
try:
    with sqlite3.connect("BaseDBicicletas.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS UNIDADES \
                            (Clave INTEGER PRIMARY KEY, \
                            Rodada INTEGER NOT NULL, \
                            Color TEXT NOT NULL);")
except Error as e:
    print(e)
except Exception:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

#Tabla Clientes
try:
    with sqlite3.connect("BaseDBicicletas.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS CLIENTES \
                            (Clave INTEGER PRIMARY KEY, \
                            Apellidos TEXT NOT NULL, \
                            Nombres TEXT NOT NULL, \
                            Telefono TEXT NOT NULL);")
except Error as e:
    print(e)
except Exception:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

#Tabla Prestamos
try:
    with sqlite3.connect("BaseDBicicletas.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS PRESTAMOS \
                            (Folio INTEGER PRIMARY KEY, ClaveUnidad INTEGER NOT NULL, \
                            ClaveCliente INTEGER, \
                            FechaPrestamo TIMESTAP NOT NULL, \
                            DiasPrestamo INTEGER NOT NULL, \
                            FechaRetorno TIMESTAP NULL, \
                            FOREIGN KEY(ClaveUnidad) REFERENCES UNIDADES(Clave), \
                            FOREIGN KEY(ClaveCliente) REFERENCES CLIENTES(Clave));")
except Error as e:
    print(e)
except Exception:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        
def mostrar_colores():
    print()
    for clave,color in colores.items():
        print(f"{clave}. {color}")
    print()

def validar_color(color): 
    if color in colores: 
        color_validado = colores[color] 
        return color_validado 
    else: 
        print("\n***ERROR INGRESE UNA OPCION VALIDA***")
        return None

def validar_texto(variable):
    if not variable.replace(' ', '').isalpha():
        print("El texto debe contener solo letras")
        return False
    if len(variable) <= 40 and len(variable) > 0:
        return True
    elif len(variable) == 0:
        print("No se puede dejar el campo vacío")
        return False
    else:
        print("Excede los 40 caracteres permitidos")
        return False


def validar_telefono(telefono):
    if len(telefono) == 10 and telefono.isdigit():
        return True
    else:
        print("El teléfono debe tener exactamente 10 dígitos numéricos.")
        return False


#validación de claves en prestamo
def validar_clave(clave, tabla):
    if not clave.isdigit():
        print("La clave debe ser un numero.")
        return False
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:            
            mi_cursor = conn.cursor()
            mi_cursor.execute(f"SELECT Clave FROM {tabla} WHERE Clave={clave}")
            registros = mi_cursor.fetchall()
            if registros:
                return True
            else:
                print("No se encontro la clave")
                return False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    return False


def mostrar_reporte_tabulado(headers, datos):
    print(tabulate(datos, headers=headers, tablefmt="grid"))

def mostrar_unidades_existentes(formato=None, imprimir=True):
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM UNIDADES")
            unidades_existentes = mi_cursor.fetchall()
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

    if imprimir:
        print("\n--- UNIDADES EN EXISTENCIA ---")
        mostrar_reporte_tabulado(["CLAVE", "RODADA", "COLOR"], unidades_existentes)

    if formato:
        if formato == 1:
            with open("Listado_de_unidades_completo.csv", "w", encoding='latin1', newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("CLAVE", "RODADA", "COLOR"))
                grabador.writerows(unidades_existentes)
        elif formato == 2:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(("CLAVE", "RODADA", "COLOR"))
            for cell in ws[1]:  # La primera fila (encabezados)
                cell.font = Font(bold=True)  # Negritas
                cell.alignment = Alignment(horizontal='center')  # Texto centrado
            for fila in unidades_existentes:
                ws.append(fila)
                for cell in ws[ws.max_row]:  # Iterar sobre la última fila añadida
                    cell.alignment = Alignment(horizontal='center')  # Centrar texto en cada celda
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho
            wb.save("Listado_de_unidades_completo.xlsx")


def mostrar_listado_de_unidades(listado_de_unidades=None,formato=None):
    if listado_de_unidades == 2: #unidades por rodada
        try:
            with sqlite3.connect("BaseDBicicletas.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT Clave, Color FROM UNIDADES WHERE Rodada = (?)", (rodada_ingresada,))
                unidades_por_rodada = mi_cursor.fetchall()
        except Error as e:
            print(e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close()
        if unidades_por_rodada:
            print(f"\n--- UNIDADES CON RODADA {rodada_ingresada} ---")
            mostrar_reporte_tabulado(["CLAVE", "COLOR"], unidades_por_rodada)

            if formato:
                if formato == 1:
                    with open("Listado_de_unidades_xRodada.csv", "w", encoding='latin1', newline="") as archivo:
                        grabador = csv.writer(archivo)
                        grabador.writerow(("CLAVE", "COLOR"))
                        grabador.writerows(unidades_por_rodada)
                elif formato == 2:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(("CLAVE", "COLOR"))
                    for cell in ws[1]:  # La primera fila (encabezados)
                        cell.font = Font(bold=True)  # Negritas
                        cell.alignment = Alignment(horizontal='center')  # Texro centrado
                    for fila in unidades_por_rodada:
                        ws.append(fila)
                        for cell in ws[ws.max_row]:  # Iterar sobre la última fila añadida
                            cell.alignment = Alignment(horizontal='center')  # Centrar texto en cada celda
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter  # Obtener la letra de la columna
                        for cell in col:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho
                    wb.save("Listado_de_unidades_xRodada.xlsx")
        else:
            print(f"No se encontraron unidades con rodada {rodada_ingresada}.")


def mostrar_clientes_registrados():
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT Clave, Apellidos, Nombres FROM CLIENTES")
            clientes_registrados = mi_cursor.fetchall()
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    print("\n--- Clientes ---")
    mostrar_reporte_tabulado(["CLAVE", "APELLIDOS", "NOMBRES"], clientes_registrados)

def generar_reporte_clientes(formato=None):
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM CLIENTES")
            reporte_clientes = mi_cursor.fetchall()
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    print("\n--- Reporte de Clientes ---")
    mostrar_reporte_tabulado(["CLAVE", "APELLIDOS", "NOMBRES", "TELÉFONO"], reporte_clientes)
    
    if formato:
        if formato == 1:
            with open("Reporte_Clientes.csv", "w", encoding='latin1', newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("CLAVE", "APELLIDOS", "NOMBRES", "TELEFONO"))
                grabador.writerows(reporte_clientes)
        elif formato == 2:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(("CLAVE", "APELLIDOS", "NOMBRES", "TELEFONO"))
            for cell in ws[1]:
                cell.font = Font(bold=True)  # Negritas
                cell.alignment = Alignment(horizontal='center')  # Texro centrado
            for fila in reporte_clientes:
                ws.append(fila)
                for cell in ws[ws.max_row]:  # Iterar sobre la última fila añadida
                    cell.alignment = Alignment(horizontal='center')  # Centrar texto en cada celda
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho
            wb.save("Reporte_Clientes.xlsx")

def info_cliente_especifico(clave_cliente_reporte):
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:            
            mi_cursor = conn.cursor()
            mi_cursor.execute(f"SELECT * FROM CLIENTES WHERE Clave={clave_cliente_reporte}")
            info_cliente = mi_cursor.fetchall()
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    for clave, nombre, apellido, telefono in info_cliente:
        print(f"\nClave: {clave}\nApellido(s): {apellido}\nNombre(s): {nombre}\nTelefono: {telefono}\n")  

def generar_reporte_cliente_especifico(clave_cliente_reporte, formato=None, imprimir=True):
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de las tablas desde la base de datos
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
            df_clientes = pd.read_sql_query("SELECT * FROM CLIENTES", conn)
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)
        
        # Convertir 'FechaPrestamo' a datetime y luego a solo fechas
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo'])

        df_prestamos['FechaDebioRetorno'] = df_prestamos['FechaPrestamo'] + pd.to_timedelta(df_prestamos['DiasPrestamo'], unit='D')
        
        # Filtrar los préstamos con el cliente proporcionado
        df_prestamos_filtrado = df_prestamos[(df_prestamos['ClaveCliente'] == clave_cliente_reporte) ]
        
        if df_prestamos_filtrado.empty:
            print("No se encontraron préstamos para el cliente especificado.")
            return  # Detener si no hay datos en el rango de fechas
        
        # Unir los DataFrames de unidades y préstamos
        df_merged = df_prestamos_filtrado.merge(df_unidades, left_on='ClaveUnidad', right_on='Clave')
        df_merged = df_merged.merge(df_clientes, left_on='ClaveCliente', right_on='Clave')
        
        # Concatenar nombre y apellidos en una sola columna
        df_merged['NombreCompleto'] = df_merged['Nombres'] + ' ' + df_merged['Apellidos']
        
        # Seleccionar y ordenar columnas
        columnas_interes = ['ClaveUnidad', 'Rodada', 'Color','FechaPrestamo','FechaDebioRetorno', 'FechaRetorno']
        df_reporte = df_merged[columnas_interes].sort_values(by='FechaPrestamo')
        
        # Formatear las fechas según el formato deseado
        df_reporte['FechaPrestamo'] = df_reporte['FechaPrestamo'].apply(lambda x: x.strftime("%m-%d-%Y"))
        df_reporte['FechaDebioRetorno'] = df_reporte['FechaDebioRetorno'].apply(lambda x: x.strftime("%m-%d-%Y"))
        
        # Mostrar el reporte solo si imprimir es True
        if imprimir:
            #print(df_reporte)
            print(f"\n--- Historial de prestamos del cliente: {clave_cliente_reporte} ---")
            mostrar_reporte_tabulado(["CLAVE UNIDAD", "RODADA", "COLOR", "FECHA PRESTAMO", "FECHA DEBIDO","FECHA RETORNO"], df_reporte)
    
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
        return
    except Exception as e:
        print(f"Error: {e}")
        return

    # Exportar el reporte en formato CSV o Excel
    if formato == 1:
        with open(f"Reporte Prestamos Cliente #{clave_cliente_reporte}.csv", "w", encoding='latin1', newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(('CLAVE DE UNIDAD', 'TAMANO DE RODADA','COLOR', 'FECHA DE PRESTAMO', 'FECHA DEBIDO','FECHA RETORNO'))
            grabador.writerows(df_reporte.values)
    elif formato == 2:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(('CLAVE DE UNIDAD', 'TAMANO DE RODADA', 'COLOR','FECHA DE PRESTAMO', 'FECHA DEBIDO','FECHA RETORNO'))
        for cell in ws[1]:  # La primera fila (encabezados)
            cell.font = Font(bold=True)  # Negritas
            cell.alignment = Alignment(horizontal='center')  # Texto centrado

        # Agregar filas
        for fila in df_reporte.values:
            ws.append(list(fila))  # Pasar cada fila como lista de valores

        # Ajustar el formato de las columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

        # Guardar el archivo de Excel
        wb.save(f"Reporte Prestamos Cliente #{clave_cliente_reporte}.xlsx")


def mostrar_prestamos_por_retornar(formato=None):
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT Folio, ClaveUnidad, ClaveCliente, FechaPrestamo, DiasPrestamo FROM PRESTAMOS WHERE FechaRetorno IS NULL")
            prestamos_por_retornar = mi_cursor.fetchall()
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    if prestamos_por_retornar:
         print("\n--- Préstamos por Retornar ---")
         mostrar_reporte_tabulado(["FOLIO", "CLAVE UNIDAD", "CLAVE CLIENTE", "FECHA PRESTAMO", "DIAS PRESTAMO"], prestamos_por_retornar)
    

def generar_reporte_prestamos_por_retornar(fecha_inicio, fecha_fin, formato=None, imprimir=True):
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de las tablas desde la base de datos
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
            df_clientes = pd.read_sql_query("SELECT * FROM CLIENTES", conn)
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)
        
        # Convertir 'FechaPrestamo' a datetime y luego a solo fechas
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo']).dt.date
        
        # Filtrar los préstamos dentro del rango de fechas proporcionado y que no tengan fecha de entrega
        df_prestamos_filtrado = df_prestamos[
            (df_prestamos['FechaPrestamo'] >= fecha_inicio) & 
            (df_prestamos['FechaPrestamo'] <= fecha_fin) & 
            (df_prestamos['FechaRetorno'].isnull())  # Filtrar préstamos sin fecha de entrega
        ]
        
        if df_prestamos_filtrado.empty:
            print("No se encontraron préstamos por retornar en el rango de fechas especificado.")
            return
        
        # Unir los DataFrames de unidades y préstamos
        df_merged = df_prestamos_filtrado.merge(df_unidades, left_on='ClaveUnidad', right_on='Clave')
        df_merged = df_merged.merge(df_clientes, left_on='ClaveCliente', right_on='Clave')
        
        # Concatenar nombre y apellidos en una sola columna
        df_merged['NombreCompleto'] = df_merged['Nombres'] + ' ' + df_merged['Apellidos']
        
        # Seleccionar y ordenar columnas
        columnas_interes = ['ClaveUnidad', 'Rodada', 'FechaPrestamo', 'NombreCompleto', 'Telefono']
        df_reporte = df_merged[columnas_interes].sort_values(by='FechaPrestamo')
        
        # Formatear las fechas según el formato deseado
        df_reporte['FechaPrestamo'] = df_reporte['FechaPrestamo'].apply(lambda x: x.strftime("%m-%d-%Y"))
        
        # Mostrar el reporte solo si imprimir es True
        if imprimir:
            #print(df_reporte)
            print("\n--- PRESTAMOS POR RETORNAR ---")
            mostrar_reporte_tabulado(["CLAVE UNIDAD", "RODADA", "FECHA PRESTAMO", "NOMBRE COMPLETO", "TELEFONO"], df_reporte)

    except FileNotFoundError as e:
        print(f"El archivo no fue encontrado: {e.filename}. Asegúrate de que el archivo esté en la ruta correcta.")
        return
    except Exception as e:
        print(f"Error: {e}")
        return

    # Exportar el reporte en formato CSV o Excel
    if formato == 1:
        with open("Reporte_Prestamos_Por_Retornar.csv", "w", encoding='latin1', newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(('CLAVE DE UNIDAD', 'TAMANO DE RODADA', 'FECHA DE PRESTAMO', 'NOMBRE COMPLETO', 'TELEFONO'))
            grabador.writerows(df_reporte.values)
    elif formato == 2:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(('CLAVE DE UNIDAD', 'TAMANO DE RODADA', 'FECHA DE PRESTAMO', 'NOMBRE COMPLETO', 'TELEFONO'))
        for cell in ws[1]:  # La primera fila (encabezados)
            cell.font = Font(bold=True)  # Negritas
            cell.alignment = Alignment(horizontal='center')  # Texto centrado

        # Agregar filas
        for fila in df_reporte.values:
            ws.append(list(fila))  # Pasar cada fila como lista de valores

        # Ajustar el formato de las columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

        # Guardar el archivo de Excel
        wb.save("Reporte_Prestamos_Por_Retornar.xlsx")


def generar_reporte_prestamos_por_periodo(fecha_inicio, fecha_fin, formato=None, imprimir=True):
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de las tablas desde la base de datos
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
            df_clientes = pd.read_sql_query("SELECT * FROM CLIENTES", conn)
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)
        
        # Convertir 'FechaPrestamo' a datetime y luego a solo fechas
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo']).dt.date
        
        # Filtrar los préstamos dentro del rango de fechas proporcionado
        df_prestamos_filtrado = df_prestamos[(df_prestamos['FechaPrestamo'] >= fecha_inicio) & (df_prestamos['FechaPrestamo'] <= fecha_fin)]
        
        if df_prestamos_filtrado.empty:
            print("No se encontraron préstamos en el rango de fechas especificado.")
            return  # Detener si no hay datos en el rango de fechas
        
        # Unir los DataFrames de unidades y préstamos
        df_merged = df_prestamos_filtrado.merge(df_unidades, left_on='ClaveUnidad', right_on='Clave')
        df_merged = df_merged.merge(df_clientes, left_on='ClaveCliente', right_on='Clave')
        
        # Concatenar nombre y apellidos en una sola columna
        df_merged['NombreCompleto'] = df_merged['Nombres'] + ' ' + df_merged['Apellidos']
        
        # Seleccionar y ordenar columnas
        columnas_interes = ['ClaveUnidad', 'Rodada', 'FechaPrestamo', 'NombreCompleto', 'Telefono']
        df_reporte = df_merged[columnas_interes].sort_values(by='FechaPrestamo')
        
        # Formatear las fechas según el formato deseado
        df_reporte['FechaPrestamo'] = df_reporte['FechaPrestamo'].apply(lambda x: x.strftime("%m-%d-%Y"))
        
        # Mostrar el reporte solo si imprimir es True
        if imprimir:
            #print(df_reporte)
            print("\n--- PRESTAMOS POR PERIODO ---")
            mostrar_reporte_tabulado(["CLAVE UNIDAD", "RODADA", "FECHA PRESTAMO", "NOMBRE COMPLETO", "TELEFONO"], df_reporte)
    
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
        return
    except Exception as e:
        print(f"Error: {e}")
        return

    # Exportar el reporte en formato CSV o Excel
    if formato == 1:
        with open("Reporte_Prestamos_Por_Periodo.csv", "w", encoding='latin1', newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(('CLAVE DE UNIDAD', 'TAMANO DE RODADA', 'FECHA DE PRESTAMO', 'NOMBRE COMPLETO', 'TELEFONO'))
            grabador.writerows(df_reporte.values)
    elif formato == 2:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(('CLAVE DE UNIDAD', 'TAMANO DE RODADA', 'FECHA DE PRESTAMO', 'NOMBRE COMPLETO', 'TELEFONO'))
        for cell in ws[1]:  # La primera fila (encabezados)
            cell.font = Font(bold=True)  # Negritas
            cell.alignment = Alignment(horizontal='center')  # Texto centrado

        # Agregar filas
        for fila in df_reporte.values:
            ws.append(list(fila))  # Pasar cada fila como lista de valores

        # Ajustar el formato de las columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

        # Guardar el archivo de Excel
        wb.save("Reporte_Prestamos_Por_Periodo.xlsx")

def analizar_prestamos():
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de la tabla de préstamos
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)

        # Asegurarte de que la columna "DÍAS DE PRÉSTAMO" existe
        if 'DiasPrestamo' in df_prestamos.columns:
            # Calcular los estadísticos manualmente
            mean = df_prestamos["DiasPrestamo"].mean()
            median = df_prestamos["DiasPrestamo"].median()
            mode = df_prestamos["DiasPrestamo"].mode()[0]  # Moda podría tener múltiples valores, tomamos el primero
            min_value = df_prestamos["DiasPrestamo"].min()
            max_value = df_prestamos["DiasPrestamo"].max()
            std_dev = df_prestamos["DiasPrestamo"].std()
            # Calcular cuartiles (percentiles 25, 50, 75)
            q1 = df_prestamos["DiasPrestamo"].quantile(0.25)
            q2 = median  # El cuartil 2 (Q2) es la mediana
            q3 = df_prestamos["DiasPrestamo"].quantile(0.75)
            # Mostrar los estadísticos con un índice personalizado
            print("\nEstadísticas descriptivas de 'DÍAS DE PRÉSTAMO':")
            estadisticos = pd.Series([mean, median, mode, min_value, max_value, std_dev, q1, q2, q3], 
                                     index=["Media", "Mediana", "Moda", "Mínimo", "Máximo", "Desviación Estándar", 
                                            "Cuartil 1 (Q1)", "Cuartil 2 (Q2)", "Cuartil 3 (Q3)"])
        else:
            print("La columna 'DiasPrestamo' no se encontró en la base de datos.")
    except FileNotFoundError:
        print("La base de datos 'BaseDBicicletas.db' no fue encontrada. Asegúrate de que esté en la ruta correcta.")
    except Exception as e:
        print(f"Error: {e}")

def validar_unidad_prestada(clave_unidad):
    conflicto = False
    with sqlite3.connect("BaseDBicicletas.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT * FROM PRESTAMOS WHERE ClaveUnidad = ? AND FechaRetorno IS NULL", (clave_unidad,))
        prestamo_conflicto = mi_cursor.fetchall()
        if prestamo_conflicto:
            conflicto = True
    return conflicto

import pandas as pd
import sqlite3
from datetime import datetime

def reporte_retrasos(formato=None):
    fecha_actual = datetime.now()
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de las tablas desde la base de datos
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
            df_clientes = pd.read_sql_query("SELECT * FROM CLIENTES", conn)
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)

        # Convertir fechas a datetime
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo'], errors='coerce')
        df_prestamos['FechaRetorno'] = pd.to_datetime(df_prestamos['FechaRetorno'], errors='coerce')

        # Calcular la fecha en la que se debió retornar el préstamo
        df_prestamos['FechaDebioRetorno'] = df_prestamos['FechaPrestamo'] + pd.to_timedelta(df_prestamos['DiasPrestamo'], unit='D')

        # Filtrar préstamos con retraso
        df_retrasos = df_prestamos[(df_prestamos['FechaRetorno'].isna()) & (df_prestamos['FechaDebioRetorno'] < fecha_actual)].copy()

        if df_retrasos.empty:
            print("No hay préstamos retrasados.")
            return

        # Calcular días de retraso
        df_retrasos.loc[:, 'DiasRetraso'] = (fecha_actual - df_retrasos['FechaDebioRetorno']).dt.days

        # Unir DataFrames
        df_merged = df_retrasos.merge(df_unidades, left_on='ClaveUnidad', right_on='Clave', how='left')
        df_merged = df_merged.merge(df_clientes, left_on='ClaveCliente', right_on='Clave', how='left')

        # Concatenar nombre y apellidos en una sola columna
        df_merged['NombreCompleto'] = df_merged['Nombres'] + ' ' + df_merged['Apellidos']

        # Seleccionar y ordenar columnas de interés
        columnas_interes = ['DiasRetraso', 'FechaDebioRetorno', 'ClaveUnidad', 'Rodada', 'Color', 'NombreCompleto', 'Telefono']
        df_reporte = df_merged[columnas_interes].sort_values(by='DiasRetraso', ascending=False)

        # Formatear las fechas según el formato deseado
        df_reporte['FechaDebioRetorno'] = df_reporte['FechaDebioRetorno'].apply(lambda x: x.strftime("%m-%d-%Y"))

        # Mostrar el reporte
        print("\n--- PRESTAMOS CON RETRASO ---")
        mostrar_reporte_tabulado(["DIAS DE RETRASO","FECHA DEBIDO","CLAVE UNIDAD", "RODADA", "COLOR", "NOMBRE COMPLETO", "TELEFONO"], df_reporte)


        # Exportar el reporte en formato CSV o Excel
        if formato == 1:
            with open("Reporte_Retrasos.csv", "w", encoding='latin1', newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("DIAS DE RETRASO","FECHA DEBIDO","CLAVE UNIDAD", "RODADA", "COLOR", "NOMBRE COMPLETO", "TELEFONO"))
                grabador.writerows(df_reporte.values)
        elif formato == 2:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(("DIAS DE RETRASO","FECHA DEBIDO","CLAVE UNIDAD", "RODADA", "COLOR", "NOMBRE COMPLETO", "TELEFONO"))
            for cell in ws[1]:  # La primera fila (encabezados)
                cell.font = Font(bold=True)  # Negritas
                cell.alignment = Alignment(horizontal='center')  # Texto centrado
            # Agregar filas
            for fila in df_reporte.values:
                ws.append(list(fila))  # Pasar cada fila como lista de valores
            # Ajustar el formato de las columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho
            # Guardar el archivo de Excel
            wb.save("Reporte_Retrasos.xlsx")
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
        return
    except Exception as e:
        print(f"Error: {e}")


def ranking_clientes():
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de las tablas desde la base de datos
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
            df_clientes = pd.read_sql_query("SELECT * FROM CLIENTES", conn)
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)
        
        # Convertir 'FechaPrestamo' a datetime y luego a solo fechas
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo']).dt.date
        
        # Unir los DataFrames de unidades y préstamos
        df_merged = df_prestamos.merge(df_unidades, left_on='ClaveUnidad', right_on='Clave')
        df_merged = df_merged.merge(df_clientes, left_on='ClaveCliente', right_on='Clave')
        
        # Concatenar nombre y apellidos en una sola columna
        df_merged['NombreCompleto'] = df_merged['Nombres'] + ' ' + df_merged['Apellidos']
        
        # Crear el ranking de clientes
        ranking = df_merged.groupby('ClaveCliente').agg({
            'FechaPrestamo': 'count',
            'NombreCompleto': 'first',
            'Telefono': 'first'
        }).reset_index()
        
        # Renombrar columnas
        ranking = ranking.rename(columns={'FechaPrestamo': 'CantidadRentas'})
        
        # Ordenar por cantidad de rentas acumuladas de manera descendente
        ranking = ranking.sort_values(by='CantidadRentas', ascending=False)
        
        # Seleccionar columnas para mostrar
        ranking = ranking[['CantidadRentas', 'ClaveCliente', 'NombreCompleto', 'Telefono']]
        
        # Mostrar el ranking
        print(ranking)
    
    except Exception as e:
        print(f"Error: {e}")

def cantidad_prestamos_xRodada():
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de la tabla UNIDADES
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
        
        # Contar préstamos por tamaño de rodada y ordenar
        contador_prestamos_xRodada = df_unidades.groupby('Rodada').size().reset_index(name="cantidad_prestamos").sort_values(by="cantidad_prestamos", ascending=False)
        
        # Mostrar el resultado
        print(contador_prestamos_xRodada)

        # Crear gráfica de pastel
        plt.figure(figsize=(8, 8))
        plt.pie(contador_prestamos_xRodada['cantidad_prestamos'], labels=contador_prestamos_xRodada['Rodada'], autopct='%1.1f%%')
        plt.title('Proporción de Préstamos por Rodada')
        plt.show()
    
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")


def cantidad_prestamos_xColor():
    try:
        # Conectarse a la base de datos
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            # Leer los datos de la tabla UNIDADES
            df_unidades = pd.read_sql_query("SELECT * FROM UNIDADES", conn)
        
        # Contar préstamos por color y ordenar
        contador_prestamos_xColor = df_unidades.groupby('Color').size().reset_index(name="cantidad_prestamos").sort_values(by="cantidad_prestamos", ascending=False)
        
        # Mostrar el resultado
        print(contador_prestamos_xColor)
        colores_asignados = {'Rojo': '#ff0000', 'Verde': '#11e120', 'Azul': '#2016f0', 'Amarillo': '#f3ff00',\
         'Naranja': '#ff6c00', 'Morado': '#760dd3'} 
        colores = [colores_asignados[color] for color in contador_prestamos_xColor['Color']]
        # Crear gráfica de pastel
        plt.figure(figsize=(8, 8))
        plt.pie(contador_prestamos_xColor['cantidad_prestamos'], labels=contador_prestamos_xColor['Color'], \
        autopct='%1.1f%%',colors=colores)
        plt.title('Proporción de Préstamos por Color')
        plt.show()
        
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")

def cantidad_prestamos_xDiaSemana():
    try:
        with sqlite3.connect("BaseDBicicletas.db") as conn:
            df_prestamos = pd.read_sql_query("SELECT * FROM PRESTAMOS", conn)
        
        df_prestamos['FechaPrestamo'] = pd.to_datetime(df_prestamos['FechaPrestamo'])
        df_prestamos['DiaSemana'] = df_prestamos['FechaPrestamo'].dt.dayofweek
        
        # Mapeo para cambiar el inicio de la semana a domingo (0) y ajustar los días correspondientes
        mapeo_dias = {0: 'Domingo', 1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes', 6: 'Sábado'}
        df_prestamos['DiaSemana'] = df_prestamos['DiaSemana'].map(mapeo_dias)
        
        contador_prestamos_xDiaSemana = df_prestamos.groupby('DiaSemana').size().reset_index(name="cantidad_prestamos")
        
        orden_dias = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        contador_prestamos_xDiaSemana['DiaSemana'] = pd.Categorical(contador_prestamos_xDiaSemana['DiaSemana'], categories=orden_dias, ordered=True)
        contador_prestamos_xDiaSemana = contador_prestamos_xDiaSemana.sort_values('DiaSemana')

        print("\nCantidad de préstamos totales acumulados por día de la semana\n")
        print(contador_prestamos_xDiaSemana)


        # Crear gráfica de barras
        plt.figure(figsize=(10, 6))
        plt.bar(contador_prestamos_xDiaSemana['DiaSemana'], contador_prestamos_xDiaSemana['cantidad_prestamos'], color='skyblue')
        plt.xlabel('Día de la Semana')
        plt.ylabel('Cantidad de Préstamos')
        plt.title('Cantidad de Préstamos Totales Acumulados por Día de la Semana')
        plt.show()

    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")

def pregunta_exportar():
    while True:
        try:
            exportar = int(input("¿Desea exportar el reporte? (1.-si/2.-no): "))
            if exportar == 1: return exportar
            elif exportar == 2:
                print("Saliendo...")
                return None
        except ValueError:
            print("ERROR ***Ingrese una opción válida***")

def pregunta_formato():
    while True:
        try:
            formato = int(input("¿En qué formato desea exportar el reporte (1.-csv/2.-excel)?: "))
            if formato==1 or formato==2:return formato
        except ValueError:
            print("ERROR: ***Ingrese una opción válida***")

while True:
    try:
        print("\n***MENU PRINCIPAL***")
        menu=int(input("\n1. Registro\n2. Prestamo\n3. Retorno\n4. Informes\n5. Salir\nIngrese el numero de opcion que desee: "))
        
        if menu == 1:
            while True:
                try:
                    print("\nMENU PRINCIPAL > REGISTRO")
                    opcion_registro = int(input("\n1. Unidad\n2. Cliente\n3. Salir\nSeleccione la opcion que desea registrar: "))
                    
                    if opcion_registro == 1:
                        while True:
                            print("\nMENU PRINCIPAL > REGISTRO > UNIDAD")
                            try:
                                rodada = int(input("\nEscoge el tamaño de la rodada 20, 26 o 29\nIngrese el numero: "))
                                if rodada == 20 or rodada == 26 or rodada == 29:
                                    break
                                else:
                                    print("***Ingrese un tamaño de rodada valido***\n")
                            except ValueError:
                                print("***INGRESE UNA OPCION VALIDA***")

                        while True:
                            try:
                                mostrar_colores()
                                op_color = int(input("Ingrese la opción de color que desee: "))
                                color = validar_color(op_color)
                                if color:
                                    break
                            except ValueError:
                                print("***INGRESE UNA OPCIÓN VÁLIDA***")
                                
                        try:
                            with sqlite3.connect("BaseDBicicletas.db") as conn:
                                mi_cursor = conn.cursor()
                                unidades_insert = {"Rodada": rodada, "Color": color}
                                mi_cursor.execute("INSERT INTO UNIDADES (Rodada, Color) VALUES(:Rodada, :Color)", unidades_insert)
                        except Error as e:
                            print(e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()

                    elif opcion_registro == 2:
                        while True:
                            print("\nMENU PRINCIPAL > REGISTRO > CLIENTE")
                            apellidos = input("\nIngrese los apellidos del cliente: ")
                            if validar_texto(apellidos): break
                        while True:
                            nombres = input("Ingrese los nombres del cliente: ")
                            if validar_texto(nombres): break
                        while True:
                            telefono = input("Ingrese el teléfono del cliente: ")
                            if validar_telefono(telefono): break
                        try:
                            with sqlite3.connect("BaseDBicicletas.db") as conn:
                                mi_cursor = conn.cursor()
                                clientes_insert={"Apellido":apellidos, "Nombre":nombres, "Telefono":telefono}
                                mi_cursor.execute("INSERT INTO CLIENTES (Apellidos, Nombres, Telefono) \
                                                  VALUES(:Apellido, :Nombre, :Telefono)", clientes_insert)
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                    elif opcion_registro == 3:
                        break
                except ValueError:
                    print("***ERROR INGRESE UNA OPCION VALIDA***")

        elif menu == 2:
            while True:
                conflicto = False  
                registro_exitoso = False
                salir_del_bucle = False
                folio = max(prestamos, default=0) + 1
                print("\nMENU PRINCIPAL > PRESTAMO")
                mostrar_unidades_existentes()
                while True:
                    clave_unidad = input("\n***Presione ENTER si desea salir***\nIngrese la clave de la unidad: ")
                    if clave_unidad.strip() == "":
                        salir_del_bucle = True
                        print("Saliendo...")
                        break
                    if validar_clave(clave_unidad, tabla_unidad):
                        if not validar_unidad_prestada(clave_unidad):
                            mostrar_clientes_registrados()

                            while True:
                                clave_cliente = input("Ingrese la clave del cliente: ")
                                if validar_clave(clave_cliente, tabla_cliente):
                                    while True:
                                        try:
                                            fecha_str = input("\n***Presione ENTER si desea registrar la fecha actual***\nIngrese la fecha en formato mm-dd-aaaa: ")
                                            if fecha_str.strip() == "":
                                                fecha_prestamo = datetime.today().strftime("%m-%d-%Y")
                                                print(f"Fecha del préstamo registrada: {fecha_prestamo}")
                                                break
                                            else:
                                                fecha_prestamo_manual = datetime.strptime(fecha_str, "%m-%d-%Y").date()
                                                if fecha_prestamo_manual >= datetime.today().date():
                                                    fecha_prestamo = fecha_prestamo_manual.strftime("%m-%d-%Y")
                                                    print(f"Fecha del préstamo registrada: {fecha_prestamo}")
                                                    break
                                                else:
                                                    print("El día ingresado no puede ser anterior a la fecha actual.")
                                        except ValueError:
                                            print("***Ingrese la fecha en formato mm-dd-aaaa***")

                                    # Lógica de captura de días de préstamo
                                    while True:
                                        try:
                                            dias_prestamo = int(input("\nIngrese la cantidad de días del préstamo: "))
                                            if 0 < dias_prestamo <= 14:
                                                break
                                            else:
                                                print("La cantidad de días debe ser mayor a cero y menor a 14.")
                                        except ValueError:
                                            print("Ingrese un número válido por favor.")

                                    fecha_retorno = None
                                    try:
                                        with sqlite3.connect("BaseDBicicletas.db") as conn:
                                            mi_cursor = conn.cursor()
                                            prestamos_insert = {"ClaveUnidad": clave_unidad, "ClaveCliente": clave_cliente,
                                                                "FechaPrestamo": fecha_prestamo, "DiasPrestamo": dias_prestamo,
                                                                "FechaRetorno": fecha_retorno}
                                            mi_cursor.execute("INSERT INTO PRESTAMOS (ClaveUnidad, ClaveCliente, FechaPrestamo, \
                                                            DiasPrestamo, FechaRetorno)\
                                                            VALUES(:ClaveUnidad, :ClaveCliente, :FechaPrestamo, \
                                                            :DiasPrestamo, :FechaRetorno)", prestamos_insert)
                                        registro_exitoso = True   
                                    except Error as e:
                                        print(e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    finally:
                                        conn.close()
                                    salir_del_bucle = True
                                    break
                        else:
                            print("***ERROR: LA UNIDAD YA ESTÁ PRESTADA, ELIJA UNA DISPONIBLE***") 
                    if salir_del_bucle:
                        break
                if salir_del_bucle:
                    break

        elif menu == 3:
            print("\nMENU PRINCIPAL > RETORNO")
            mostrar_prestamos_por_retornar()
            try:
                while True:
                    folio_input = input("\n***Presione ENTER si desea salir***\nIngrese el folio del préstamo para registrar la devolución: ")
                    if folio_input.replace(" ", "") == "":
                        print("Saliendo...")
                        break
                    try:
                        folio = int(folio_input)
                        with sqlite3.connect("BaseDBicicletas.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute(f"SELECT Folio, FechaPrestamo FROM PRESTAMOS WHERE Folio = {folio}")
                            registros = mi_cursor.fetchall()
                            mi_cursor.execute(f"SELECT Folio, FechaPrestamo FROM PRESTAMOS WHERE FechaRetorno = NOT NULL")
                            retornado = mi_cursor.fetchall
                            if not registros:
                                print(f"El folio no existe o la unidad ya fue devuelta")
                            elif retornado:
                                print(f"El prestamo ya fue retornado")

                            for folio_select, fecha_prestamo_select in registros:
                                folio = folio_select
                                fecha_prestamo = fecha_prestamo_select
                            if isinstance(fecha_prestamo, str):
                                fecha_prestamo = datetime.strptime(fecha_prestamo, "%m-%d-%Y").date()
                    except ValueError:
                        print("Ingrese un número válido.")
                        continue
                    except Error as e:
                        print(f"Error en la base de datos: {e}")
                        continue
                    except Exception:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        continue

                    while True:
                        try:
                            fecha_retorno = input("Ingrese la fecha de retorno en formato mm-dd-yyyy: ")
                            try:
                                fecha_retorno_validada = datetime.strptime(fecha_retorno, "%m-%d-%Y").date()
                                if fecha_retorno_validada >= fecha_prestamo:
                                    try:
                                        with sqlite3.connect("BaseDBicicletas.db") as conn:
                                            mi_cursor = conn.cursor()
                                            update_retorno = {"FechaRetorno": fecha_retorno_validada.strftime("%m-%d-%Y")}
                                            mi_cursor.execute(f"UPDATE PRESTAMOS SET FechaRetorno = :FechaRetorno WHERE Folio = {folio}", update_retorno)
                                            break
                                    except Error as e:
                                        print(f"Error actualizando el registro: {e}")
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                else:
                                    print("La fecha de retorno debe ser igual o posterior a la fecha de préstamo.")
                            except ValueError:
                                print("Error: Ingrese una fecha válida en formato mm-dd-yyyy")
                        except ValueError:
                            print("Error: Ingrese una fecha válida en formato mm-dd-yyyy")
                    break
            except ValueError:
                print("ERROR: ***Ingrese una opción válida***")

        elif menu == 4:
            while True:
                try:
                    print("\nMENU PRINCIPAL > INFORMES")
                    opcion_informes = int(input("\n1. Reportes\n2. Análisis\n3. Salir\nSeleccione la opción que desea: "))

                    if opcion_informes == 1:
                        while True:
                            try:
                                print("\nMENU PRINCIPAL > INFORMES > REPORTES\n")
                                opcion_reporte = int(input("1. Clientes\n2. Listado de unidades\n3. Retrasos\n4. Préstamos por Retornar\n5. Préstamos por Periodo\n6. Salir\nSeleccione el tipo de reporte: "))

                                if opcion_reporte == 1:
                                    while True:
                                        try:
                                            print("\nMENU PRINCIPAL > INFORMES > REPORTES > CLIENTES\n")
                                            opcion_reporte_clientes = int(input("1. Reporte completo de clientes\n2. Reporte de cliente especifico\n3. Salir\n Seleccione el tipo de reporte de clientes: "))
                                            if opcion_reporte_clientes == 1:
                                                print("\nMENU PRINCIPAL > INFORMES > REPORTES > CLIENTES > COMPLETO\n")
                                                generar_reporte_clientes()
                                                while True:
                                                    exportar=pregunta_exportar()
                                                    if exportar:
                                                        formato=pregunta_formato()
                                                        generar_reporte_clientes(formato)
                                                        break
                                                    break
                                            elif opcion_reporte_clientes == 2:
                                                print("\nMENU PRINCIPAL > INFORMES > REPORTES > CLIENTES > ESPECIFICO\n")
                                                mostrar_clientes_registrados()
                                                while True:
                                                    clave_cliente_reporte = input("Ingrese la clave del cliente del cual desea realizar un reporte: ")
                                                    if validar_clave(clave_cliente_reporte, tabla_cliente):
                                                        clave_cliente_reporte=int(clave_cliente_reporte)
                                                        info_cliente_especifico(clave_cliente_reporte)
                                                        generar_reporte_cliente_especifico(clave_cliente_reporte)
                                                        while True:
                                                            exportar=pregunta_exportar()
                                                            if exportar:
                                                                formato=pregunta_formato()
                                                                generar_reporte_cliente_especifico(clave_cliente_reporte, formato, imprimir=False)
                                                                break
                                                            break
                                                        break

                                            elif opcion_reporte_clientes == 3:
                                                print("Saliendo...")
                                                break
                                        except ValueError:
                                            print("***ERROR INGRESE UNA OPCION VALIDA***")

                                elif opcion_reporte == 2:
                                    while True:
                                        try:
                                            print("\nMENU PRINCIPAL > INFORMES > REPORTES > LISTADO DE UNIDADES >\n")
                                            listado_de_unidades = int(input("1. Listado Completo\n2. Listado por Rodada\n3. Listado por Color\n4. Volver al menú\nElija una opcion: "))

                                            if listado_de_unidades == 1:
                                                print("\nMENU PRINCIPAL > INFORMES > REPORTES > LISTADO DE UNIDADES > COMPLETO\n")
                                                mostrar_unidades_existentes()  # Mostrar inicialmente
                                                while True:
                                                    exportar=pregunta_exportar()
                                                    if exportar:
                                                        formato=pregunta_formato()
                                                        mostrar_unidades_existentes(formato, imprimir=False) 
                                                        break
                                                    break


                                            elif listado_de_unidades == 2:
                                                print("\nMENU PRINCIPAL > INFORMES > REPORTES > LISTADO DE UNIDADES > xRODADA")
                                                while True:
                                                    try:
                                                        rodada_ingresada = int(input("\nIngrese la rodada que desea filtrar: "))
                                                        if rodada_ingresada==20 or rodada_ingresada==26 or rodada_ingresada==29:
                                                            break
                                                        else:
                                                            print("***ERROR INGRESE UNA RODADA EXISTENTE***")
                                                    except ValueError:
                                                        print("***ERROR INGRESE UNA OPCION VALIDA*** :(")
                                                mostrar_listado_de_unidades(listado_de_unidades)
                                                while True:
                                                    exportar=pregunta_exportar()
                                                    if exportar:
                                                        formato=pregunta_formato()
                                                        mostrar_listado_de_unidades(listado_de_unidades, formato)
                                                        break
                                                    break
                                            
                                            elif listado_de_unidades == 3:
                                                print("\nMENU PRINCIPAL > INFORMES > REPORTES > LISTADO DE UNIDADES > xCOLOR")
                                                while True:
                                                    while True:
                                                        try:
                                                            mostrar_colores()
                                                            opcion_color = int(input("Ingrese la opcion de color que desea filtrar: "))
                                                            color_ingresado = validar_color(opcion_color)
                                                            if color_ingresado: break
                                                        except ValueError:
                                                            print("\n***ERROR INGRESE UNA OPCION VALIDA***")
                                                    try:
                                                        with sqlite3.connect("BaseDBicicletas.db") as conn:
                                                            mi_cursor = conn.cursor()
                                                            mi_cursor.execute(f"SELECT Clave, Rodada FROM UNIDADES WHERE Color = (?)", (color_ingresado,))
                                                            unidades_por_color = mi_cursor.fetchall()
                                                    except Error as e:
                                                        print(e)
                                                        continue
                                                    except Exception:
                                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                        continue
                                                    if unidades_por_color:
                                                        print(f"\n--- UNIDADES EN COLOR {color_ingresado} ---")
                                                        mostrar_reporte_tabulado(["CLAVE", "RODADA"], unidades_por_color)
                                                        while True:
                                                            exportar=pregunta_exportar()
                                                            if exportar:
                                                                formato=pregunta_formato()
                                                                if formato == 1:
                                                                    with open("Listado_de_unidades_xColor.csv", "w", encoding='latin1', newline="") as archivo:
                                                                        grabador = csv.writer(archivo)
                                                                        grabador.writerow(("CLAVE", "RODADA"))
                                                                        grabador.writerows(unidades_por_color)
                                                                elif formato == 2:
                                                                    wb = openpyxl.Workbook()
                                                                    ws = wb.active
                                                                    ws.append(("CLAVE", "RODADA"))
                                                                    for cell in ws[1]:  # La primera fila (encabezados)
                                                                        cell.font = Font(bold=True)  # Negritas
                                                                        cell.alignment = Alignment(horizontal='center')  # Texto centrado
                                                                    for fila in unidades_por_color:
                                                                        ws.append(fila)
                                                                        for cell in ws[ws.max_row]:  # Iterar sobre la última fila añadida
                                                                            cell.alignment = Alignment(horizontal='center')  # Centrar texto en cada celda
                                                                    for col in ws.columns:
                                                                        max_length = 0
                                                                        col_letter = col[0].column_letter  # Obtener la letra de la columna
                                                                        for cell in col:
                                                                            try:
                                                                                max_length = max(max_length, len(str(cell.value)))
                                                                            except:
                                                                                pass
                                                                        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho
                                                                    wb.save("Listado_de_unidades_xColor.xlsx")
                                                                break
                                                            break
                                                        break
                                                    else:
                                                        print(f"No se encontraron unidades con el color {color_ingresado}.")
                                                        
                                            elif listado_de_unidades == 4:
                                                print("Saliendo...")
                                                break

                                            else:
                                                print("***Ingrese una opcion valida")
                                        except ValueError:
                                            print("***ERROR INGRESE UNA OPCION VALIDA***")

                                elif opcion_reporte==3:
                                    print("\nMENU PRINCIPAL > INFORMES > REPORTES > RETRASOS")
                                    reporte_retrasos()
                                    while True:
                                        exportar=pregunta_exportar()
                                        if exportar:
                                            formato=pregunta_formato()
                                            reporte_retrasos(formato)
                                            break
                                        break
                                            
                                elif opcion_reporte == 4:
                                    print("\nMENU PRINCIPAL > INFORMES > REPORTES > xRETORNAR\n")
                                    while True:
                                        fecha_inicio = input("Ingrese la fecha de inicio (mm-dd-aaaa): ")
                                        try:
                                            # Convertir la fecha a objeto datetime.date
                                            fecha_inicio_validada = datetime.strptime(fecha_inicio, "%m-%d-%Y").date()
                                            break
                                        except ValueError:
                                            print("Ingrese la fecha en formato mm-dd-aaaa")

                                    while True:
                                        fecha_fin = input("Ingrese la fecha de fin (mm-dd-aaaa): ")
                                        try:
                                            # Convertir la fecha a objeto datetime.date
                                            fecha_fin_validada = datetime.strptime(fecha_fin, "%m-%d-%Y").date()
                                            break
                                        except ValueError:
                                            print("Ingrese la fecha en formato mm-dd-aaaa")

                                    # Mostrar el reporte inicialmente
                                    generar_reporte_prestamos_por_retornar(fecha_inicio_validada, fecha_fin_validada)

                                    while True:
                                        exportar=pregunta_exportar()
                                        if exportar:
                                            formato=pregunta_formato()
                                            generar_reporte_prestamos_por_retornar(fecha_inicio_validada, fecha_fin_validada, formato, imprimir=False)
                                            break
                                        break



                                elif opcion_reporte == 5:
                                    print("\nMENU PRINCIPAL > INFORMES > REPORTES > xPERIODO\n")
                                    while True:
                                        fecha_inicio = input("Ingrese la fecha de inicio (mm-dd-aaaa): ")
                                        try:
                                            fecha_inicio_validada = datetime.strptime(fecha_inicio, "%m-%d-%Y").date()
                                            break
                                        except ValueError:
                                            print("Ingrese la fecha en formato mm-dd-aaaa")

                                    while True:
                                        fecha_fin = input("Ingrese la fecha de fin (mm-dd-aaaa): ")
                                        try:
                                            fecha_fin_validada = datetime.strptime(fecha_fin, "%m-%d-%Y").date()
                                            break
                                        except ValueError:
                                            print("Ingrese la fecha en formato mm-dd-aaaa")

                                    # Mostrar el reporte inicialmente
                                    generar_reporte_prestamos_por_periodo(fecha_inicio_validada, fecha_fin_validada)

                                    while True:
                                        exportar=pregunta_exportar()
                                        if exportar:
                                            formato=pregunta_formato()
                                            generar_reporte_prestamos_por_periodo(fecha_inicio_validada, fecha_fin_validada, formato, imprimir=False)
                                            break
                                        break

                                elif opcion_reporte == 6:
                                    print("Saliendo...")
                                    break
                            except ValueError:
                                print("***ERROR INGRESE UNA OPCION VALIDA***")

                    elif opcion_informes == 2:
                        while True:
                            try:
                                print("\nMENU PRINCIPAL > INFORMES > ANALISIS \n")
                                opcion_analisis = int(input("1. Duración de los préstamos\n2. Ranking de los clientes\n3. Preferencias de rentas\n4. Volver al menú de informes\nSeleccione el tipo de Análisis: "))

                                if opcion_analisis == 1:
                                    print("\nMENU PRINCIPAL > INFORMES > ANALISIS  > DURACION DE PRESTAMOS")
                                    analizar_prestamos()

                                elif opcion_analisis == 2:
                                    print("\nMENU PRINCIPAL > INFORMES > ANALISIS  > RANKING DE CLIENTES\n")
                                    ranking_clientes()

                                elif opcion_analisis == 3:
                                    
                                    while True:
                                        print("\nMENU PRINCIPAL > INFORMES > ANALISIS > PREFERENCIAS DE RENTAS")
                                        print("\nPreferencias de rentas")
                                        print("1. Por rodada")
                                        print("2. Por Color")
                                        print("3. Por Día de la semana")
                                        print("4. Salir")
                                        opcion = input("Elija una opción: ")

                                        if opcion == "1":
                                            print("\nMENU PRINCIPAL > INFORMES > ANALISIS > PREFERENCIAS DE RENTAS > xRODADA\n")
                                            cantidad_prestamos_xRodada()
                                        elif opcion == "2":
                                            print("\nMENU PRINCIPAL INFORMES > ANALISIS > PREFERENCIAS DE RENTAS > xCOLOR\n")
                                            cantidad_prestamos_xColor()
                                        elif opcion == "3":
                                            print("\nMENU PRINCIPAL > INFORMES > ANALISIS > PREFERENCIAS DE RENTAS > xDIA DE LA SEMANA")
                                            cantidad_prestamos_xDiaSemana()
                                        elif opcion == "4":
                                            print("Saliendo...")
                                            break
                                        else:
                                            print("Opción no válida. Intente de nuevo.")

                                elif opcion_analisis == 4:
                                    print("Saliendo...")
                                    break

                            except ValueError:
                                print("***ERROR INGRESE UNA OPCIÓN VÁLIDA***")
                    elif opcion_informes == 3:
                        print("Saliendo...")
                        break
                except ValueError:
                    print("***ERROR INGRESE UNA OPCIÓN VÁLIDA***")                   
        elif menu==5: break
        elif menu>5 or menu<1: print("ERROR ***Ingrese una opcion valida***")
    except ValueError:
        print("ERROR ***Ingrese una opcion valida***")
